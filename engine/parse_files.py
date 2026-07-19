"""通用檔案解析分派：txt / csv / xlsx / pdf / jpeg。

解析結果一律回傳 {"kind", "summary", "data", "needs_review"}，
由 Web 介面呈現供承辦人確認後才寫入（human-in-the-loop）。
"""

from __future__ import annotations

import csv
import io
import os
import re
from datetime import date

from .parse_excel import MinistryExcel, NotMinistryExcel

TEXT_EXT = {".txt", ".md"}
SHEET_EXT = {".xlsx", ".xlsm", ".xls"}
CSV_EXT = {".csv", ".tsv"}
PDF_EXT = {".pdf"}
IMG_EXT = {".jpg", ".jpeg", ".png"}

SUPPORTED = TEXT_EXT | SHEET_EXT | CSV_EXT | PDF_EXT | IMG_EXT


# ── LINE 文字擷取（規則式，LLM 為後續強化）────────────────
_RE_THEME = re.compile(r"年度主題[：:]\s*(.+)")
_RE_RECRUIT = re.compile(r"招募\s*(\d+)\s*位?弟兄[、,\s]*(\d+)\s*位?姊妹")
_RE_BIBLE_TARGET = re.compile(r"贈經目標[：:]\s*([\d,]+)\s*本")
_RE_PERIOD = re.compile(r"(20\d{2})\s*[-–]\s*(20\d{2})")
_RE_SCHED = re.compile(r"(\d{1,2})/(\d{1,2})\s*(?:\(([一二三四五六日])\))?\s*(\d{1,2}:\d{2})?\s*[：:]\s*(.+)")
_RE_ROTA = re.compile(r"第([一二三四五])週[：:]\s*([^\s(（]+)\s*[（(]?([^)）]*)[)）]?")

_CN_NUM = {"一": 1, "二": 2, "三": 3, "四": 4, "五": 5}


MAX_SCHEDULE = 40
# LINE 匯出的日期分隔行：「2023.10.24 星期二」
_RE_DAY_HEADER = re.compile(r"^(20\d{2})[./-](\d{1,2})[./-](\d{1,2})\s")
MAX_TARGET_LEN = 40
_GIVING_HINT = ("贈經", "醫院", "診所", "安養", "機構", "教會", "學校",
                "國中", "高中", "大學", "國小", "旅館", "監獄", "分送")
# LINE 匯出的訊息行開頭：「上午11:23\t姓名\t內容」
_RE_LINE_MSG = re.compile(r"^(上午|下午)?\s*\d{1,2}:\d{2}\t")


def _strip_line_export(text: str) -> str:
    """去掉 LINE 匯出的時間戳與說話者欄位，只留訊息內容。"""
    out = []
    for raw in text.splitlines():
        if _RE_LINE_MSG.match(raw):
            parts = raw.split("\t")
            out.append(parts[-1] if len(parts) >= 3 else raw)
        else:
            out.append(raw)
    return "\n".join(out)


# 訊息行：「13:37<sep>說話者<sep>內容」，分隔可能是 tab 或空白
_RE_MSG = re.compile(r"^(?:(上午|下午)\s*)?(\d{1,2}:\d{2})[\t ]+(\S{1,20})[\t ]+(.*)$")


def _split_msg(raw: str) -> tuple[str, str, str]:
    """回傳 (時間, 說話者, 內容)；多行訊息的續行只有內容。"""
    line = raw.rstrip()
    if m := _RE_MSG.match(line):
        ampm, t, who, text = m.groups()
        return (f"{ampm or ''}{t}", who, text.strip())
    return ("", "", line.strip())


def split_by_day(text: str) -> list[tuple[str, list[str]]]:
    """依 LINE 日期分隔行切段，回傳 [(YYYY-MM-DD, [訊息行…]), …]。"""
    days: list[tuple[str, list[str]]] = []
    cur: list[str] | None = None
    for raw in text.splitlines():
        if m := _RE_DAY_HEADER.match(raw.strip()):
            iso = f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
            cur = []
            days.append((iso, cur))
        elif cur is not None and raw.strip():
            cur.append(raw)
    return days


def filter_month(text: str, year: int, month: int) -> dict:
    """從整份聊天記錄篩出報告月份的訊息（原始匯出動輒跨數年）。"""
    prefix = f"{year}-{month:02d}"
    days = split_by_day(text)
    hit = [(d, lines) for d, lines in days if d.startswith(prefix)]
    msgs = []
    for d, lines in hit:
        last_who = ""
        for raw in lines:
            time, who, text = _split_msg(raw)
            if who:
                last_who = who
            msgs.append({"date": d, "time": time,
                         "who": who or last_who, "text": text,
                         "continued": not who})
    return {"total_days": len(days),
            "range": (days[0][0], days[-1][0]) if days else None,
            "matched_days": len(hit), "messages": msgs}


def parse_line_text(text: str, report_year: int) -> dict:
    """從 LINE 匯出文字擷取年度主題、贈經排程、禱告會輪值。

    LINE 原始匯出是整份聊天記錄（可能上萬行、含多年舊訊息），
    故採「保守擷取」：只取結構明確的條列，去重、限長、限量，
    結果一律交由介面人工確認後才寫入。
    """
    text = _strip_line_export(text)
    out: dict = {}

    # ── 年度主題：僅在有「年度主題」字樣的段落附近取值 ────
    annual: dict = {}
    if m := _RE_THEME.search(text):
        annual["theme"] = m.group(1).strip()[:60]
        window = text[max(0, m.start() - 300): m.start() + 600]
        if p := _RE_PERIOD.search(window):
            annual["period"] = f"{p.group(1)}-{p.group(2)}"
        if r := _RE_RECRUIT.search(window):
            annual["recruit_target"] = {"brothers": int(r.group(1)),
                                        "sisters": int(r.group(2))}
        if b := _RE_BIBLE_TARGET.search(window):
            annual["bible_target"] = int(b.group(1).replace(",", ""))
    if annual:
        out["annual"] = annual

    # ── 贈經排程：條列格式 MM/DD(週) HH:MM：地點 ──────────
    schedule, seen = [], set()
    for raw in text.splitlines():
        line = raw.strip().lstrip("•・-‧*　 ")
        if not line or len(line) > 120:
            continue
        m = _RE_SCHED.search(line)
        if not m:
            continue
        mm, dd, wd, tm, target = m.groups()
        target = target.strip()
        # 過濾聊天長句，只留短且具贈經語意的目標
        if len(target) > MAX_TARGET_LEN or not any(k in target or k in line
                                                   for k in _GIVING_HINT):
            continue
        try:
            iso = date(report_year, int(mm), int(dd)).isoformat()
        except ValueError:
            continue
        key = (iso, target)
        if key in seen:
            continue
        seen.add(key)
        schedule.append({"date": iso, "weekday": wd or "", "time": tm or "",
                         "target": target, "status": "已排定"})
    if schedule:
        out["bible_giving"] = {"schedule": schedule[:MAX_SCHEDULE],
                               "truncated": len(schedule) > MAX_SCHEDULE}

    # ── 禱告會輪值 ────────────────────────────────────
    weeks, wseen = [], set()
    for m in _RE_ROTA.finditer(text):
        w = _CN_NUM.get(m.group(1), 0)
        if w in wseen:
            continue
        wseen.add(w)
        weeks.append({"week": w, "place": m.group(2).strip()[:12],
                      "time": m.group(3).strip()[:20]})
    if len(weeks) >= 3:   # 需成組出現才視為輪值表
        out["prayer_rota"] = {"weeks": sorted(weeks, key=lambda x: x["week"])}

    return out


# ── 分派 ─────────────────────────────────────────────────
def parse_file(path: str, module: str = "", report_year: int | None = None,
               report_month: int | None = None) -> dict:
    ext = os.path.splitext(path)[1].lower()
    name = os.path.basename(path)
    report_year = report_year or date.today().year
    report_month = report_month or date.today().month

    if ext in SHEET_EXT:
        try:
            mx = MinistryExcel(path)
            a = mx.analysis()
            return {"kind": "ministry_excel", "file": name,
                    "summary": f"事工成果表（{a['date_range'] or '—'}）："
                               f"{a['people']} 位會員、已繳會費 {a['fee_paid']}、"
                               f"教會見證 {a['church_count']} 間、"
                               f"贈經 {a['official']['scripture_total']:,} 本",
                    "data": mx.to_model(), "needs_review": False}
        except NotMinistryExcel:
            pass
        except Exception as exc:
            return {"kind": "sheet_error", "file": name,
                    "summary": f"讀取失敗：{exc}", "data": {}, "needs_review": True}
        return _parse_generic_sheet(path, name)

    if ext in CSV_EXT:
        with open(path, newline="", encoding="utf-8-sig", errors="replace") as f:
            rows = list(csv.reader(f, delimiter="\t" if ext == ".tsv" else ","))
        return {"kind": "csv", "file": name,
                "summary": f"CSV：{len(rows)} 列 × {len(rows[0]) if rows else 0} 欄",
                "data": {"rows": rows[:200]}, "needs_review": True}

    if ext in TEXT_EXT:
        text = open(path, encoding="utf-8", errors="replace").read()
        month = report_month
        extracted = parse_line_text(text, report_year)
        bits = []
        if "annual" in extracted:
            bits.append("年度主題與目標")
        if "bible_giving" in extracted:
            bits.append(f"贈經排程 {len(extracted['bible_giving']['schedule'])} 筆")
        if "prayer_rota" in extracted:
            bits.append(f"禱告會輪值 {len(extracted['prayer_rota']['weeks'])} 週")

        data = {"extracted": extracted, "raw_preview": text[:3000],
                "lines": text.count("\n") + 1}
        if month:
            fm = filter_month(text, report_year, month)
            data["month_digest"] = {**fm, "messages": fm["messages"][:200]}
            span = f"（記錄涵蓋 {fm['range'][0]} ~ {fm['range'][1]}）" if fm["range"] else ""
            bits.append(f"{report_year}年{month}月訊息 {len(fm['messages'])} 則{span}")

        return {"kind": "line_text", "file": name,
                "summary": ("擷取到：" + "、".join(bits)) if bits else
                           "未偵測到結構化欄位（原始聊天記錄），已保留原文供人工挑選",
                "data": data, "needs_review": True}

    if ext in PDF_EXT:
        return _parse_pdf(path, name)

    if ext in IMG_EXT:
        return {"kind": "image", "file": name,
                "summary": "影像已收檔，需以多模態辨識（代禱會/贈經照片）",
                "data": {"path": path},
                "needs_review": True, "needs_ocr": True}

    return {"kind": "unsupported", "file": name,
            "summary": f"不支援的格式 {ext}", "data": {}, "needs_review": True}


def _parse_generic_sheet(path: str, name: str) -> dict:
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    sheets = {}
    for ws in wb.worksheets[:3]:
        sheets[ws.title] = [
            [c if not hasattr(c, "isoformat") else c.isoformat() for c in row]
            for row in ws.iter_rows(max_row=60, max_col=30, values_only=True)
        ]
    return {"kind": "sheet", "file": name,
            "summary": f"試算表：{len(wb.worksheets)} 個工作表（{', '.join(wb.sheetnames[:5])}）",
            "data": {"sheets": sheets}, "needs_review": True}


def _parse_pdf(path: str, name: str) -> dict:
    try:
        import pdfplumber
    except ImportError:
        return {"kind": "pdf", "file": name,
                "summary": "未安裝 pdfplumber，無法抽取文字",
                "data": {}, "needs_review": True}
    text, tables = [], []
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages[:20]:
            text.append(page.extract_text() or "")
            for t in page.extract_tables() or []:
                tables.append(t)
    joined = "\n".join(text)
    scanned = len(joined.strip()) < 50
    return {"kind": "pdf", "file": name,
            "summary": ("掃描型 PDF（無文字層），需 OCR" if scanned
                        else f"PDF：{len(text)} 頁、{len(tables)} 個表格"),
            "data": {"text": joined[:8000], "tables": tables[:10]},
            "needs_review": True, "needs_ocr": scanned}
