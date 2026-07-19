"""由 Grafana API 資料重建「支會各項事工成果表」Excel。

用途：新財年的年度目標尚未建檔時，`/report/campstat` 會回 500 產不出報表，
此模組直接用 API 資料重建同樣版面，讓既有解析流程與人工檢視都能沿用。

版面完全比照官方匯出檔（已用 2026/06 實際匯出檔核對）：
  Row 1 標題與區間｜Row 2 列印日期｜Row 3-4 欄位標頭
  Row 5 目標｜Row 6 成果｜Row 7 差額｜Row 8 達成率｜Row 9+ 各欄區塊
  A 編號｜D 弟兄 E 姊妹｜F-G 會費｜H-I 聖經奉獻｜J-K 巴拿巴
  L 見證日期 M 講員 N 教會 O 奉獻｜P 贈經類別 Q 本數

⚠️ 各欄區塊是**各自獨立的清單**，不是同一列對應同一人
   （官方檔即如此：教會見證那四欄與左邊的會員無關）。
"""

from __future__ import annotations

from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side

from .grafana import MINISTRY_CATEGORIES

SHEET = "支會各項事工成果表"

# Row 5-8 匯總列：欄索引 → category
SUMMARY_LAYOUT = {
    4: "增加弟兄", 5: "增加姊妹",
    6: "弟兄會費", 7: "姊妹會費",
    8: "弟兄聖奉", 9: "姊妹聖奉",
    12: "教會見證", 15: "教會聖奉",
    16: "贈送聖經", 17: "姊妹贈經",
}

THIN = Side(style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _num(v):
    return v if isinstance(v, (int, float)) else None


def build(data: dict, dest_path: str, team: str = "海山") -> dict:
    """data 為 grafana.fetch_all() 的回傳值。回傳寫檔摘要。"""
    fy = data["fiscal_year"]
    stats = data["ministry_stats"]
    roster = data.get("member_roster") or []
    by_member = data.get("offerings_by_member") or {}
    churches = data.get("church_testimony") or []
    bibles = data.get("bible_giving") or []

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET

    # ── Row 1-2 標題 ────────────────────────────────────
    ws.cell(1, 7, f"{team}支會各項事工成果表").font = Font(size=14, bold=True)
    ws.cell(1, 12, f" ({fy - 1}/06/01-{fy}/05/31 )")
    ws.cell(2, 10, "不列入八項成果目標")
    ws.cell(2, 12, "入帳截止日期：")
    ws.cell(2, 16, "列印日期")
    ws.cell(2, 17, date.today().strftime("%m/%d"))

    # ── Row 3-4 欄位標頭 ────────────────────────────────
    ws.cell(3, 4, "弟兄"); ws.cell(3, 5, "姊妹")
    ws.cell(3, 6, "會員會費")
    ws.cell(3, 8, "聖經奉獻 ")
    ws.cell(3, 10, "巴拿巴奉獻\n(總會行政經費奉獻)")
    ws.cell(3, 12, "教會見證奉獻")
    ws.cell(3, 16, "贈經")

    ws.cell(4, 1, "現有"); ws.cell(4, 2, "弟兄"); ws.cell(4, 3, "姊妹")
    ws.cell(4, 4, sum(1 for m in roster if m["brother"]))
    ws.cell(4, 5, sum(1 for m in roster if m["sister"]))
    for col, label in ((6, "弟兄"), (7, "姊妹"), (8, "弟兄"), (9, "姊妹"),
                       (10, "弟兄"), (11, "姊妹"), (12, "日期"), (13, "講員"),
                       (14, "名稱"), (15, "奉獻"), (16, "贈經總數"), (17, "姊妹贈經")):
        ws.cell(4, col, label)

    # ── Row 5-8 匯總 ────────────────────────────────────
    for row, (label, field) in enumerate(
            (("目標", "goal"), ("成果", "value"),
             ("差額", "diff"), ("達成率", "rate")), start=5):
        ws.cell(row, 1, label)
        for col, category in SUMMARY_LAYOUT.items():
            rec = stats.get(category)
            ws.cell(row, col, rec.get(field) if rec else None)

    # ── Row 9+ 各欄區塊（彼此獨立）───────────────────────
    r = 9
    for m in roster:
        ws.cell(r, 1, m["id"])
        ws.cell(r, 4, m["brother"])
        ws.cell(r, 5, m["sister"])
        ws.cell(r, 6, m["fee_brother"])
        ws.cell(r, 7, m["fee_sister"])
        key = str(m["id"])
        ws.cell(r, 8, _num(by_member.get("弟兄聖奉", {}).get(key)))
        ws.cell(r, 9, _num(by_member.get("姊妹聖奉", {}).get(key + "A")))
        ws.cell(r, 10, _num(by_member.get("巴拿巴", {}).get(key)))
        ws.cell(r, 11, _num(by_member.get("巴拿巴", {}).get(key + "A")))
        r += 1
    last_member_row = r - 1

    for i, c in enumerate(churches):
        rr = 9 + i
        ws.cell(rr, 12, c.get("date"))
        ws.cell(rr, 13, c.get("speaker"))
        ws.cell(rr, 14, c.get("church"))
        ws.cell(rr, 15, _num(c.get("amount")) or None)

    # 贈經：依類型彙總（官方檔 P/Q 欄即為類別＋本數）
    kinds: dict[str, float] = {}
    for b in bibles:
        kinds[b["kind"]] = kinds.get(b["kind"], 0) + (b["count"] or 0)
    for i, (kind, n) in enumerate(sorted(kinds.items(), key=lambda x: -x[1])):
        ws.cell(9 + i, 16, kind)
        ws.cell(9 + i, 17, n)

    # ── 樣式 ────────────────────────────────────────────
    for row in ws.iter_rows(min_row=3, max_row=max(last_member_row, 12),
                            min_col=1, max_col=17):
        for cell in row:
            cell.border = BORDER
            if cell.row <= 8:
                cell.alignment = CENTER
    for row in (5, 6, 7):
        for col in (8, 9, 10, 11, 15, 16):
            ws.cell(row, col).number_format = "#,##0"
    for col in (8, 9, 10, 11, 15):
        ws.cell(8, col).number_format = "0%"
    for col, w in ((1, 9), (4, 11), (5, 11), (6, 9), (7, 9), (8, 10), (9, 10),
                   (10, 10), (11, 10), (12, 11), (13, 11), (14, 24), (15, 11),
                   (16, 12), (17, 11)):
        ws.column_dimensions[chr(64 + col)].width = w

    wb.save(dest_path)
    return {
        "path": dest_path, "sheet": SHEET, "fiscal_year": fy,
        "period": f"{fy - 1}/06/01-{fy}/05/31",
        "members": len(roster), "churches": len(churches),
        "bible_kinds": len(kinds),
        "note": "由 Grafana API 重建（官方 Excel 報表於本財年無法產生）",
    }
