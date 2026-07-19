"""年度贈經計畫（聖經配送計畫）解析。

Excel 版面：Sheet `配送計畫` 把歷年區塊**並排**在同一張表上，
第 1 列是各年度的標題（如「2026-2027年度學校(2026/6/1-2027/5/31)贈經計畫規劃」），
第 2 列是該區塊的欄位標頭，資料由第 3 列往下。

因此不能硬編欄號——每年新增一個區塊，欄位就往右移。
本模組掃第 1 列找出目標財年的起始欄，再依第 2 列標頭定位各欄。

已知資料瑕疵（實測 2026-2027）：
  - 中正國中的「時間」是 `1900-01-05`，Excel 時間格式錯亂，實際應為 15:30
  - 「鶯歌工商」「柑園國中」只有校名、無編號與日期，屬候補場次
  - 6 月畢典多為「06月　日」，日期尚未定案
這些一律標記後照樣回傳，交由介面呈現給承辦人判斷，不自行猜測。
"""

from __future__ import annotations

import datetime as _dt
import re
from pathlib import Path

import openpyxl

SHEET = "配送計畫"
HEADER_ROW = 2
FIRST_DATA_ROW = 3

# 第 2 列可能出現的欄位標頭 → 內部欄名
HEADERS = {
    "月份": "month_label", "No": "no", "贈經日": "date", "贈經日期": "date",
    "項目": "kind", "學校": "school", "星期": "weekday",
    "時間": "time", "贈經數": "count",
}

_RE_MD = re.compile(r"(\d{1,2})\s*月\s*(\d{1,2})?\s*日?")


def _norm_time(v) -> str:
    """時間正規化。Excel 把部分儲存格存成 datetime，屬格式錯亂。"""
    if v is None:
        return ""
    if isinstance(v, _dt.time):
        return v.strftime("%H:%M")
    if isinstance(v, _dt.datetime):
        # 1900-01-0X 這類是 Excel 的時間序列錯亂，不是真實日期
        return "" if v.year <= 1900 else v.strftime("%H:%M")
    return str(v).strip()


def find_period_column(ws, period: str) -> int | None:
    """掃第 1 列找出該財年區塊的起始欄（回傳該區塊最左欄）。"""
    want = period.replace("-", "")
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v and want in str(v).replace("-", "").replace(" ", ""):
            return c
    return None


def _column_map(ws, start_col: int, width: int = 10) -> dict[str, int]:
    """由第 2 列的標頭建立 {欄名: 欄號}。往左多看一欄以涵蓋「月份」。"""
    out: dict[str, int] = {}
    for c in range(max(1, start_col - 1), start_col + width):
        h = ws.cell(HEADER_ROW, c).value
        if not h:
            continue
        key = HEADERS.get(str(h).strip())
        if key and key not in out:
            out[key] = c
    return out


def parse(path: str | Path, period: str) -> dict:
    """解析指定財年的配送排程。

    回傳 {period, schools: [...], warnings: [...]}；
    每筆 school 含 no/month/date/kind/school/weekday/time/count/status。
    """
    wb = openpyxl.load_workbook(str(path), data_only=True)
    if SHEET not in wb.sheetnames:
        return {"period": period, "schools": [], "sheet": None,
                "warnings": [f"找不到工作表「{SHEET}」，實際有：{wb.sheetnames[:5]}"]}

    ws = wb[SHEET]
    start = find_period_column(ws, period)
    if start is None:
        years = [str(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)
                 if ws.cell(1, c).value]
        return {"period": period, "schools": [], "sheet": SHEET,
                "warnings": [f"這份檔案沒有 {period} 年度的區塊。"
                             f"現有年度：{'、'.join(y[:9] for y in years)}"]}

    cols = _column_map(ws, start)
    if "school" not in cols:
        return {"period": period, "schools": [], "sheet": SHEET,
                "warnings": [f"{period} 區塊找不到「學校」欄，版面可能已變更"]}

    fy_end = int(period.split("-")[1])
    schools, warnings = [], []
    cur_month = None

    for r in range(FIRST_DATA_ROW, ws.max_row + 1):
        get = lambda k: ws.cell(r, cols[k]).value if k in cols else None  # noqa: E731

        if (ml := get("month_label")):
            cur_month = str(ml).strip()

        school = get("school")
        if not school or not str(school).strip():
            continue
        school = str(school).strip()

        raw_date = get("date")
        month = day = None
        if raw_date:
            if isinstance(raw_date, (_dt.datetime, _dt.date)):
                month, day = raw_date.month, raw_date.day
            elif (m := _RE_MD.search(str(raw_date))):
                month = int(m.group(1))
                day = int(m.group(2)) if m.group(2) else None

        # 財年 6/1 起：9~12 月屬前一個西元年
        iso = None
        if month and day:
            year = fy_end - 1 if month >= 6 else fy_end
            try:
                iso = _dt.date(year, month, day).isoformat()
            except ValueError:
                warnings.append(f"{school}：日期 {raw_date} 無法解析")

        no = get("no")
        status = "已排定"
        if not raw_date or not str(raw_date).strip():
            status = "待定"
            warnings.append(f"{school}：無贈經日期（候補或未排定）")
        elif month and not day:
            status = "日期待定"
            warnings.append(f"{school}：{raw_date} 只有月份，日期未定")

        time_raw = get("time")
        time_str = _norm_time(time_raw)
        if isinstance(time_raw, _dt.datetime) and time_raw.year <= 1900:
            warnings.append(f"{school}：時間欄為 {time_raw:%Y-%m-%d}，"
                            "Excel 格式錯亂，請人工確認（多為 15:30）")

        cnt = get("count")
        schools.append({
            "no": int(no) if isinstance(no, (int, float)) else None,
            "month_label": cur_month, "month": month,
            "date": iso, "date_raw": str(raw_date).strip() if raw_date else "",
            "kind": (str(get("kind")).strip() if get("kind") else ""),
            "school": school,
            "weekday": (str(get("weekday")).strip() if get("weekday") else ""),
            "time": time_str,
            "count": cnt if isinstance(cnt, (int, float)) else None,
            "status": status,
        })

    return {"period": period, "sheet": SHEET, "start_col": start,
            "schools": schools, "warnings": warnings}


def schools_of_month(plan: dict, year: int, month: int) -> list[dict]:
    """篩出某個月份的場次（含只有月份、日期未定者）。"""
    out = []
    for s in plan.get("schools", []):
        if s["date"] and s["date"][:7] == f"{year}-{month:02d}":
            out.append(s)
        elif not s["date"] and s["month"] == month:
            out.append(s)
    return out
