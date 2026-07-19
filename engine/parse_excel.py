"""Grafana 事工成果表 Excel 解析（Sheet：支會各項事工成果表）。

實際欄位（已用 2026/06 匯出檔驗證）：
  Row 3-4 欄位標頭｜Row 5 目標｜Row 6 成果｜Row 7 差額｜Row 8 達成率｜Row 9+ 會員
  A 編號｜D 弟兄 E 姊妹｜F-G 會費日期｜H-I 聖經奉獻｜J-K 巴拿巴
  L 見證日期 M 講員 N 教會 O 奉獻金額｜P 贈經類別 Q 贈經本數

注意：P/Q 在「會員列」是贈經明細（類別＋本數），與該欄標頭
（贈經總數／姊妹贈經）語意不同；總數一律取 Row 5-8 的匯總值。
"""

from __future__ import annotations

import datetime as _dt
import os
import re

import openpyxl

# ── 兩種版面 ────────────────────────────────────────────
# 事工成果表有兩個來源，欄位位置不同（皆已用實際檔案驗證）：
#
#   grafana  — Grafana /report/campstat 匯出，17 欄，B/C 空白
#   gideons  — gideons.tw 會員入口「匯出」鈕產生，15 欄，整體左移兩欄
#
# 版面判斷靠 Row 4 標頭文字，不硬編欄號（見 _detect_layout）。

LAYOUTS = {
    "grafana": {
        "cols": {
            "id": 1, "brother": 4, "sister": 5,
            "fee_brother": 6, "fee_sister": 7,
            "bible_brother": 8, "bible_sister": 9,
            "barnabas_brother": 10, "barnabas_sister": 11,
            "church_date": 12, "church_speaker": 13,
            "church_name": 14, "church_offering": 15,
            "scripture_kind": 16, "scripture_count": 17,
        },
        "summary": {
            "members_brother": 4, "members_sister": 5,
            "fee_brother": 6, "fee_sister": 7,
            "bible_brother": 8, "bible_sister": 9,
            "barnabas_brother": 10, "barnabas_sister": 11,
            "church_count": 12, "church_offering": 15,
            "scripture_total": 16, "scripture_sister": 17,
        },
        "current_members": {"brother": (4, 4), "sister": (4, 5)},
    },
    "gideons": {
        "cols": {
            "id": 1, "brother": 2, "sister": 3,
            "fee_brother": 4, "fee_sister": 5,
            "bible_brother": 6, "bible_sister": 7,
            "barnabas_brother": 8, "barnabas_sister": 9,
            "church_date": 10, "church_speaker": 11,
            "church_name": 12, "church_offering": 13,
            "scripture_kind": 14, "scripture_count": 15,
        },
        "summary": {
            "members_brother": 2, "members_sister": 3,
            "fee_brother": 4, "fee_sister": 5,
            "bible_brother": 6, "bible_sister": 7,
            "barnabas_brother": 8, "barnabas_sister": 9,
            "church_count": 10, "church_offering": 13,
            "scripture_total": 14, "scripture_sister": 15,
        },
        "current_members": {"brother": (4, 2), "sister": (4, 3)},
    },
}

# 向後相容：預設仍為 grafana 版
COLS = LAYOUTS["grafana"]["cols"]
SUMMARY_COLS = LAYOUTS["grafana"]["summary"]

SUMMARY_ROWS = {"target": 5, "actual": 6, "diff": 7, "rate": 8}

# 會費欄位可能出現的人工註記（非日期，不可覆寫）
FEE_NOTES = ("安息", "退會", "入會", "停權", "轉會")


def _norm(v):
    if v is None:
        return None
    if isinstance(v, (_dt.datetime, _dt.date)):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, float) and v.is_integer():
        return int(v)
    if isinstance(v, str):
        return v.strip() or None
    return v


def _num(v) -> float | int:
    """僅取數值，字串／None 視為 0。"""
    return v if isinstance(v, (int, float)) else 0


class NotMinistryExcel(ValueError):
    """此 Excel 不是事工成果表（例：現金帳、贈經計畫）。"""


class MinistryExcel:
    def __init__(self, path: str, strict: bool = True):
        self.path = path
        wb = openpyxl.load_workbook(path, data_only=True)
        named = [n for n in wb.sheetnames if "事工成果" in n]
        self.ws = wb[named[0]] if named else wb[wb.sheetnames[0]]
        self.sheet_name = self.ws.title
        if strict and not named and not self._looks_like_ministry():
            raise NotMinistryExcel(
                f"「{os.path.basename(path)}」不是事工成果表"
                f"（工作表：{', '.join(wb.sheetnames[:3])}）")
        self._parse()

    def _looks_like_ministry(self) -> bool:
        """以 Row 3-4 標頭與 Row 5-8 標籤判定，避免誤認其他 Excel。"""
        ws = self.ws
        head = " ".join(str(ws.cell(r, c).value or "")
                        for r in (1, 3, 4) for c in range(1, 18))
        labels = [str(ws.cell(r, 1).value or "") for r in range(5, 9)]
        has_head = sum(k in head for k in
                       ("會員會費", "聖經奉獻", "教會見證", "贈經", "巴拿巴")) >= 3
        has_labels = ["目標", "成果", "差額", "達成率"] == labels
        return has_head and has_labels

    def _detect_layout(self) -> str:
        """靠 Row 4 標頭定位「日期／講員／名稱」三連欄判斷版面。

        grafana 版在 L-N（12-14），gideons.tw 版在 J-L（10-12）。
        找不到就退回檢查「弟兄」在 Row 4 的位置。
        """
        ws = self.ws
        row4 = {c: str(ws.cell(4, c).value or "").strip() for c in range(1, 20)}

        for name, spec in LAYOUTS.items():
            c = spec["cols"]
            if (row4.get(c["church_date"]) == "日期"
                    and row4.get(c["church_speaker"]) == "講員"
                    and row4.get(c["church_name"]) == "名稱"):
                return name

        # 後備：Row 4 第一個「弟兄」的位置
        for c in range(1, 20):
            if row4.get(c) == "弟兄":
                return "gideons" if c <= 3 else "grafana"
        return "grafana"

    # ── 解析 ────────────────────────────────────────────
    def _parse(self):
        ws = self.ws
        self.layout = self._detect_layout()
        spec = LAYOUTS[self.layout]
        self.cols = spec["cols"]
        self.summary_cols = spec["summary"]

        # 標題與區間位置兩版不同，直接掃第 1、2 列找
        row1 = [_norm(ws.cell(1, c).value) for c in range(1, 20)]
        self.title = next((v for v in row1 if v and "成果表" in str(v)), None)
        self.date_range = next(
            (v for v in row1 if v and re.search(r"\d{4}/\d{2}/\d{2}", str(v))), None)

        cm = spec["current_members"]
        self.current_members = {
            "brother": _num(ws.cell(*cm["brother"]).value),
            "sister": _num(ws.cell(*cm["sister"]).value),
        }

        # 匯總（目標/成果/差額/達成率）
        self.summary = {
            key: {name: _norm(ws.cell(row, col).value)
                  for name, col in self.summary_cols.items()}
            for key, row in SUMMARY_ROWS.items()
        }

        # 會員列
        self.members = []
        for row in range(9, ws.max_row + 1):
            if not ws.cell(row, self.cols["id"]).value:
                continue
            rec = {k: _norm(ws.cell(row, c).value) for k, c in self.cols.items()}
            rec["row"] = row
            rec["id"] = str(rec["id"])
            self.members.append(rec)

    # ── 查詢 ────────────────────────────────────────────
    def member_by_name(self, name: str):
        name = (name or "").strip()
        if not name:
            return None
        for m in self.members:
            for f in ("brother", "sister"):
                val = m[f]
                if val and (name == str(val) or name in str(val)):
                    return m
        return None

    def churches(self) -> list[dict]:
        return [{
            "date": m["church_date"], "speaker": m["church_speaker"],
            "church": m["church_name"], "amount": _num(m["church_offering"]),
        } for m in self.members if m["church_name"]]

    def offerings(self) -> list[dict]:
        out = []
        for m in self.members:
            for gender, nk, fk, bk, kk in (
                ("M", "brother", "fee_brother", "bible_brother", "barnabas_brother"),
                ("F", "sister", "fee_sister", "bible_sister", "barnabas_sister"),
            ):
                name = m[nk]
                if not name or any(n in str(name) for n in FEE_NOTES):
                    continue
                fee = m[fk]
                fee_note = fee if fee and any(n in str(fee) for n in FEE_NOTES) else None
                out.append({
                    "id": m["id"], "name": str(name), "gender": gender,
                    "fee_date": None if fee_note else fee,
                    "fee_note": fee_note,
                    "bible_offering": _num(m[bk]),
                    "barnabas": _num(m[kk]),
                })
        return out

    def scripture_breakdown(self) -> list[dict]:
        """P/Q 欄的贈經明細（類別 → 本數）。"""
        return [{"kind": str(m["scripture_kind"]), "count": _num(m["scripture_count"])}
                for m in self.members if m["scripture_kind"]]

    # ── 分析（供 Web 檢視）───────────────────────────────
    def analysis(self) -> dict:
        offs = self.offerings()
        chs = self.churches()
        act, tgt = self.summary["actual"], self.summary["target"]
        active = [o for o in offs if not o["fee_note"]]
        paid = [o for o in active if o["fee_date"]]

        def rate(a, t):
            a, t = _num(a), _num(t)
            return round(a / t * 100, 1) if t else None

        return {
            "sheet": self.sheet_name,
            "layout": self.layout,
            "title": self.title,
            "date_range": self.date_range,
            "current_members": self.current_members,
            "member_rows": len(self.members),
            "people": len(active),
            "fee_paid": len(paid),
            "fee_unpaid": len(active) - len(paid),
            "fee_rate": round(len(paid) / len(active) * 100, 1) if active else 0,
            "bible_offering_total": sum(o["bible_offering"] for o in offs),
            "barnabas_total": sum(o["barnabas"] for o in offs),
            "church_count": len(chs),
            "church_offering_total": sum(c["amount"] for c in chs),
            # 官方匯總值（以 Excel 為準，非自行加總）
            "official": {
                "scripture_total": _num(act["scripture_total"]),
                "scripture_target": _num(tgt["scripture_total"]),
                "scripture_rate": rate(act["scripture_total"], tgt["scripture_total"]),
                "sister_scripture": _num(act["scripture_sister"]),
                "sister_target": _num(tgt["scripture_sister"]),
                "church_offering": _num(act["church_offering"]),
                "church_offering_target": _num(tgt["church_offering"]),
                "church_offering_rate": rate(act["church_offering"], tgt["church_offering"]),
                "bible_offering": _num(act["bible_brother"]) + _num(act["bible_sister"]),
                "bible_offering_target": _num(tgt["bible_brother"]) + _num(tgt["bible_sister"]),
                "recruit_actual": _num(act["members_brother"]) + _num(act["members_sister"]),
                "recruit_target": _num(tgt["members_brother"]) + _num(tgt["members_sister"]),
            },
            "summary": self.summary,
            "scripture_breakdown": self.scripture_breakdown(),
            "top_bible_offering": sorted(
                [o for o in offs if o["bible_offering"]],
                key=lambda x: x["bible_offering"], reverse=True)[:10],
            "top_churches": sorted(chs, key=lambda c: c["amount"], reverse=True)[:10],
            "unpaid_members": [o["name"] for o in active if not o["fee_date"]],
        }

    def to_ministry_stats(self) -> dict:
        """轉成與 Grafana API 相同的 {category: {goal,value,diff,rate}} 形狀。

        官方匯出檔的 Row 5 就有**年度目標**，這是 Grafana `gideons_goal1`
        缺 2027 資料時拿不到的東西，所以上傳這份檔即可補齊目標欄。
        """
        field_to_category = {
            "members_brother": "增加弟兄", "members_sister": "增加姊妹",
            "fee_brother": "弟兄會費", "fee_sister": "姊妹會費",
            "bible_brother": "弟兄聖奉", "bible_sister": "姊妹聖奉",
            "church_count": "教會見證", "church_offering": "教會聖奉",
            "scripture_total": "贈送聖經", "scripture_sister": "姊妹贈經",
        }
        out = {}
        for field, category in field_to_category.items():
            goal = _num(self.summary["target"].get(field)) or None
            val = _num(self.summary["actual"].get(field))
            if goal is None and not val:
                continue
            diff = (val - goal) if goal is not None else None
            rate = (val / goal) if goal else None
            out[category] = {"goal": goal, "value": val,
                             "diff": diff, "rate": rate}
        return out

    def to_model(self) -> dict:
        return {
            "ministry_stats": self.summary,
            "ministry_stats_api": self.to_ministry_stats(),
            "offerings": self.offerings(),
            "church_testimony": self.churches(),
            "analysis": self.analysis(),
        }
