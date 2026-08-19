"""核心邏輯單元測試：python3 -m pytest tests/ -q"""

import os
import sys
from datetime import date

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from engine.dates import (build_meta, default_report_month, fiscal_year_of,
                          fourth_sunday, roc_year, template_month)
from engine import grafana
from engine.generate import _church_key, _fmt_stat, _match_church
from engine.parse_files import filter_month, parse_line_text, split_by_day


# ── 日期／財年 ───────────────────────────────────────────
def test_roc_year():
    assert roc_year(2026) == 115


def test_fourth_sunday():
    assert fourth_sunday(2026, 7) == date(2026, 7, 26)
    assert fourth_sunday(2026, 6) == date(2026, 6, 28)


def test_template_month_wraps_january():
    assert template_month(2026, 7) == (2026, 6)
    assert template_month(2026, 1) == (2025, 12)


def test_fiscal_year_boundary():
    assert fiscal_year_of(2026, 5) == 2026    # 5 月 → 本財年
    assert fiscal_year_of(2026, 6) == 2027    # 6 月 → 下一財年
    assert fiscal_year_of(2026, 12) == 2027


def test_build_meta():
    m = build_meta(2026, 7)
    assert (m.roc_year, m.prev_month, m.fiscal_year) == (115, 6, 2027)
    assert m.meeting_date == "2026-07-26"
    assert m.period == "2026-2027"
    assert m.work_dir_name == "2026年7月月例會"
    assert "var-year=2027" in m.dashboard_url


def test_build_meta_override_meeting_date():
    assert build_meta(2026, 7, "2026-07-19").meeting_date == "2026-07-19"


# ── 達成率格式（Excel 以小數儲存，1 = 100%）──────────────
def test_fmt_rate():
    assert _fmt_stat("rate", 1, "") == "100%"
    assert _fmt_stat("rate", 0.6812, "") == "68%"
    assert _fmt_stat("actual", 184772, "") == "184,772"
    assert _fmt_stat("target", None, "人") == "-"


# ── 教會名模糊比對 ───────────────────────────────────────
def test_church_key_normalisation():
    assert _church_key("板城靈糧堂教會") == _church_key("板城靈糧堂")


def test_match_church():
    chs = [{"church": "板城靈糧堂", "amount": 13300, "date": None, "speaker": None},
           {"church": "貴格會土城教會", "amount": 5500, "date": None, "speaker": None}]
    c, how = _match_church("貴格會土城教會", chs)
    assert how == "exact" and c["amount"] == 5500
    c, how = _match_church("板城靈糧堂教會", chs)
    assert how == "fuzzy" and c["amount"] == 13300
    assert _match_church("完全不存在的教會", chs)[0] is None


def test_match_church_partial():
    """實際資料：docx 寫「樹林愛加倍教會」，DB 寫「愛加倍浸信會」。"""
    chs = [{"church": "愛加倍浸信會", "amount": 0, "date": None, "speaker": None}]
    c, how = _match_church("樹林愛加倍教會", chs)
    assert how == "partial" and c["church"] == "愛加倍浸信會"
    # 共同字串太短不應誤配
    assert _match_church("三峽真道教會", chs)[0] is None


# ── Grafana ──────────────────────────────────────────────
def test_fiscal_range():
    assert grafana.fiscal_range(2027) == ("2026-06-01", "2027-05-31")
    assert grafana.fiscal_range(2026) == ("2025-06-01", "2026-05-31")


def test_ministry_categories_cover_docx_columns():
    """-2 表 1~10 欄都要有對應的 category。"""
    assert sorted(grafana.MINISTRY_CATEGORIES) == list(range(1, 11))
    assert grafana.MINISTRY_CATEGORIES[7] == "教會聖奉"


def test_excel_fallback_off_by_default():
    """回退年度會寫入上一財年的錯誤數字，預設必須關閉。"""
    import inspect
    sig = inspect.signature(grafana.fetch_excel)
    assert sig.parameters["allow_fallback"].default is False


def test_church_testimony_includes_donation_only():
    """有奉獻但無見證場次的教會不可漏（FY2027 實際情況）。"""
    grafana.query = lambda sql, fy: [
        {"church": "基督國度溪水旁教會", "date": None, "speaker": None, "amount": 5900},
        {"church": "愛加倍浸信會", "date": "2026/07/05", "speaker": "黃哲斌", "amount": None},
    ]
    chs = grafana.church_testimony(2027)
    assert sum(c["amount"] for c in chs) == 5900
    only = [c for c in chs if c["donation_only"]]
    assert len(only) == 1 and only[0]["church"] == "基督國度溪水旁教會"
    assert [c for c in chs if c["has_event"]][0]["church"] == "愛加倍浸信會"


def test_build_excel_roundtrip(tmp_path):
    """重建的 Excel 必須能被既有解析器讀回。"""
    from engine.build_excel import build
    from engine.parse_excel import MinistryExcel
    data = {
        "fiscal_year": 2027,
        "ministry_stats": {"弟兄會費": {"goal": None, "value": 3, "diff": None, "rate": None},
                           "教會聖奉": {"goal": None, "value": 6400, "diff": None, "rate": None}},
        "member_roster": [{"id": "4412", "brother": "黃哲斌", "sister": "牛啟慧",
                           "fee_brother": "06/03", "fee_sister": "06/03",
                           "type_brother": "一般", "type_sister": "一般"}],
        "offerings_by_member": {"弟兄聖奉": {"4412": 2000}, "姊妹聖奉": {"4412A": 1000}},
        "church_testimony": [{"church": "基督國度溪水旁教會", "date": None,
                              "speaker": None, "amount": 5900}],
        "bible_giving": [],
    }
    dest = tmp_path / "out.xlsx"
    info = build(data, str(dest))
    assert info["members"] == 1 and dest.exists()

    mx = MinistryExcel(str(dest))
    assert mx.sheet_name == "支會各項事工成果表"
    assert mx.summary["actual"]["fee_brother"] == 3
    assert mx.summary["actual"]["church_offering"] == 6400
    assert mx.churches()[0]["amount"] == 5900


def _make_gideons_xlsx(path):
    """依實際觀察到的 gideons.tw 匯出版面（15 欄、整體左移兩欄）建檔。"""
    from openpyxl import Workbook
    wb = Workbook(); ws = wb.active; ws.title = "事工成果表"
    ws["E1"] = "海山支會各項事工成果表"
    ws["J1"] = " (2026/06/01-2027/05/31)"
    for col, v in zip("BCDFHJN", ["弟兄", "姊妹", "會員會費", "聖經奉獻",
                                  "巴拿巴奉獻", "教會見證奉獻", "贈經"]):
        ws[f"{col}3"] = v
    for col, v in {"A": "現有", "B": 24, "C": 16, "D": "弟兄", "E": "姊妹",
                   "F": "弟兄", "G": "姊妹", "H": "弟兄", "I": "姊妹",
                   "J": "日期", "K": "講員", "L": "名稱", "M": "奉獻",
                   "N": "贈經總數", "O": "姊妹贈經"}.items():
        ws[f"{col}4"] = v
    for r, vals in {
        5: ("目標", 3, 2, 25, 17, 55000, 15000, None, None, 10, None, None, 100000, 4000, 600),
        6: ("成果", 0, 0, 3, 2, 3000, 1000, None, None, 2, None, None, 6400, 0, None),
        7: ("差額", -3, -2, -22, -15, -52000, -14000, None, None, -8, None, None, -93600, -4000, -600),
        8: ("達成率", 0, 0, .12, .12, .05, .07, None, None, .2, None, None, .06, 0, 0),
    }.items():
        for i, v in enumerate(vals):
            ws.cell(r, i + 1, v)
    for i, m in enumerate([
        ("1192", "陳文隆", "陳雪純", "未繳", "未繳", None, None, None, None,
         "07/05", "黃哲斌", "愛加倍浸信會", None),
        ("4412", "黃哲斌", "牛啟慧", "06/03", "06/03", 2000, 1000, None, None,
         None, None, "基督國度溪水旁教會", 5900)]):
        for j, v in enumerate(m):
            ws.cell(9 + i, j + 1, v)
    wb.save(path)


def test_detect_gideons_layout(tmp_path):
    """總會匯出版面欄位左移兩欄，誤判會讀成完全錯誤的欄位。"""
    from engine.parse_excel import MinistryExcel
    p = tmp_path / "官方.xlsx"
    _make_gideons_xlsx(str(p))
    mx = MinistryExcel(str(p))
    assert mx.layout == "gideons"
    assert [m["brother"] for m in mx.members] == ["陳文隆", "黃哲斌"]
    assert mx.summary["actual"]["fee_brother"] == 3
    assert mx.summary["target"]["church_offering"] == 100000
    assert mx.churches()[1]["amount"] == 5900


def test_gideons_export_supplies_goals(tmp_path):
    """官方匯出檔的 Row 5 含年度目標——這是 Grafana 拿不到的。"""
    from engine.parse_excel import MinistryExcel
    p = tmp_path / "官方.xlsx"
    _make_gideons_xlsx(str(p))
    stats = MinistryExcel(str(p)).to_ministry_stats()
    assert stats["弟兄會費"]["goal"] == 25
    assert stats["教會聖奉"] == {"goal": 100000, "value": 6400,
                                 "diff": -93600, "rate": 0.064}


def test_grafana_layout_still_detected(tmp_path):
    """加入新版面後，原本的 17 欄版仍要正確辨識。"""
    from openpyxl import Workbook
    from engine.parse_excel import MinistryExcel
    wb = Workbook(); ws = wb.active; ws.title = "支會各項事工成果表"
    ws.cell(1, 7, "海山支會各項事工成果表")
    ws.cell(1, 12, " (2025/06/01-2026/05/31 )")
    ws.cell(3, 6, "會員會費"); ws.cell(3, 8, "聖經奉獻")
    ws.cell(3, 10, "巴拿巴奉獻"); ws.cell(3, 12, "教會見證奉獻")
    for col, v in {1: "現有", 4: 24, 5: 16, 6: "弟兄", 7: "姊妹",
                   12: "日期", 13: "講員", 14: "名稱", 15: "奉獻",
                   16: "贈經總數", 17: "姊妹贈經"}.items():
        ws.cell(4, col, v)
    for r, label in ((5, "目標"), (6, "成果"), (7, "差額"), (8, "達成率")):
        ws.cell(r, 1, label)
    ws.cell(6, 6, 24)
    ws.cell(9, 1, "1192"); ws.cell(9, 4, "陳文隆"); ws.cell(9, 5, "陳雪純")
    p = tmp_path / "grafana.xlsx"; wb.save(str(p))
    mx = MinistryExcel(str(p))
    assert mx.layout == "grafana"
    assert mx.members[0]["brother"] == "陳文隆"
    assert mx.summary["actual"]["fee_brother"] == 24


# ── 密碼保護 ─────────────────────────────────────────────
def _auth(monkeypatch, password=None, cloud=False):
    import importlib
    from engine import auth
    monkeypatch.delenv("APP_PASSWORD", raising=False)
    monkeypatch.delenv("VERCEL", raising=False)
    if password:
        monkeypatch.setenv("APP_PASSWORD", password)
    if cloud:
        monkeypatch.setenv("VERCEL", "1")
    importlib.reload(auth)
    return auth


def test_auth_off_when_local_and_unset(monkeypatch):
    a = _auth(monkeypatch)
    assert a.auth_required() is False


def test_auth_forced_on_cloud_even_without_password(monkeypatch):
    """雲端沒設密碼時必須擋下，不能無防護上線。"""
    a = _auth(monkeypatch, cloud=True)
    assert a.auth_required() is True
    assert a.configured_password() is None


def test_token_rejects_tampering(monkeypatch):
    a = _auth(monkeypatch, "correct-horse")
    good = a.issue_token()
    assert a.valid_token(good)
    exp, _, sig = good.partition(".")
    assert not a.valid_token(f"{exp}.{'0' * len(sig)}")   # 簽章竄改
    assert not a.valid_token(f"9999999999.{sig}")          # 延長效期
    assert not a.valid_token(f"1000000000.{sig}")          # 已過期
    for bad in ("", "garbage", "1.2.3", None):
        assert not a.valid_token(bad)


def test_password_check_and_strength(monkeypatch):
    a = _auth(monkeypatch, "correct-horse")
    assert a.check_password("correct-horse")
    assert not a.check_password("Correct-Horse")
    assert not a.check_password("")
    assert a.password_strength_warning() is None

    # 組織名／短密碼要被標記為弱（不在此處寫出實際使用的密碼）
    assert _auth(monkeypatch, "gideons").password_strength_warning() is not None
    assert _auth(monkeypatch, "abc123").password_strength_warning() is not None


def test_secret_changes_with_password(monkeypatch):
    """換密碼後舊 token 必須失效。"""
    a = _auth(monkeypatch, "first-password")
    old = a.issue_token()
    a = _auth(monkeypatch, "second-password")
    assert not a.valid_token(old)


# ── 儲存層：年度贈經計畫 ──────────────────────────────────
def test_local_plan_lookup_and_save(tmp_path):
    """本機版從「贈經計畫」資料夾找檔；上傳後應立刻找得到。"""
    from engine.storage import LocalStorage, plan_filename
    base, plans = tmp_path / "base", tmp_path / "plans"
    s = LocalStorage(base, [plans])
    assert s.find_plan("2026-2027") is None

    src = tmp_path / "上傳.xlsx"
    src.write_bytes(b"PK\x03\x04dummy")
    dst = s.save_plan("2026-2027", src)
    assert dst.name == plan_filename("2026-2027")
    assert s.find_plan("2026-2027") == dst
    assert s.find_plan("2027-2028") is None      # 別的年度不該誤中


def test_local_plan_dir_defaults_beside_base(tmp_path):
    from engine.storage import PLAN_FOLDER, LocalStorage
    s = LocalStorage(tmp_path / "海山支會")
    assert s.plan_dirs[0] == tmp_path / "海山支會" / PLAN_FOLDER


# ── 年度贈經計畫解析 ─────────────────────────────────────
def _make_plan_xlsx(path):
    """仿實際版面：歷年區塊並排，第 1 列年度標題、第 2 列欄位標頭。"""
    import datetime as dt

    from openpyxl import Workbook
    wb = Workbook(); ws = wb.active; ws.title = "配送計畫"
    # 舊年度區塊（欄 1-7），確保不會被誤抓
    ws.cell(1, 2, "2025-2026年度學校(2025/6/1-2026/5/31)贈經計畫規劃")
    for i, h in enumerate(["月份", "No", "贈經日期", "項目", "學校", "星期", "時間"]):
        ws.cell(2, 1 + i, h)
    ws.cell(3, 1, "9月"); ws.cell(3, 2, 1); ws.cell(3, 3, "09月18日")
    ws.cell(3, 5, "舊年度學校")

    # 目標年度區塊（欄 10 起）
    ws.cell(1, 10, "2026-2027年度學校(2026/6/1-2027/5/31)贈經計畫規劃")
    for i, h in enumerate(["月份", "No", "贈經日期", "項目", "學校", "星期", "時間", "贈經數"]):
        ws.cell(2, 10 + i, h)
    rows = [
        ("9月", 1, "09月02日", "國高中", "北大國高中", "三", dt.time(15, 30), None),
        (None, 2, "09月09日", "國高中", "安溪國中", "三", dt.time(15, 30), None),
        (None, 3, "11月04日", "國高中", "中正國中", "三", dt.datetime(1900, 1, 5), None),
        (None, 4, "01月06日", "女校", "樹人高職", "三", dt.time(15, 30), None),
        (None, None, None, None, "鶯歌工商", None, None, None),
        (None, 5, "06月　日", "國高中", "清水國高中畢典", None, dt.time(10, 45), None),
    ]
    for r, vals in enumerate(rows, start=3):
        for i, v in enumerate(vals):
            if v is not None:
                ws.cell(r, 10 + i, v)
    wb.save(path)


def test_parse_plan_picks_right_year_block(tmp_path):
    """歷年區塊並排，必須抓到指定財年而非最左邊那塊。"""
    from engine.parse_plan import parse
    p = tmp_path / "plan.xlsx"; _make_plan_xlsx(str(p))
    plan = parse(str(p), "2026-2027")
    names = [s["school"] for s in plan["schools"]]
    assert "舊年度學校" not in names
    assert "北大國高中" in names and "安溪國中" in names


def test_parse_plan_fiscal_year_date_rollover(tmp_path):
    """財年 6/1 起：9 月屬前一西元年，1 月屬後一年。"""
    from engine.parse_plan import parse
    p = tmp_path / "plan.xlsx"; _make_plan_xlsx(str(p))
    by = {s["school"]: s for s in parse(str(p), "2026-2027")["schools"]}
    assert by["北大國高中"]["date"] == "2026-09-02"
    assert by["樹人高職"]["date"] == "2027-01-06"


def test_parse_plan_flags_known_data_flaws(tmp_path):
    """時間格式錯亂、無日期候補、只有月份——都要標記而非猜測。"""
    from engine.parse_plan import parse
    p = tmp_path / "plan.xlsx"; _make_plan_xlsx(str(p))
    plan = parse(str(p), "2026-2027")
    by = {s["school"]: s for s in plan["schools"]}

    assert by["中正國中"]["time"] == ""          # 1900-01-05 不當成時間
    assert any("中正國中" in w for w in plan["warnings"])
    assert by["鶯歌工商"]["status"] == "待定"
    assert by["清水國高中畢典"]["status"] == "日期待定"


def test_parse_plan_missing_year_is_explicit(tmp_path):
    from engine.parse_plan import parse
    p = tmp_path / "plan.xlsx"; _make_plan_xlsx(str(p))
    plan = parse(str(p), "2030-2031")
    assert plan["schools"] == []
    assert "2030-2031" in plan["warnings"][0]


def test_schools_of_month(tmp_path):
    from engine.parse_plan import parse, schools_of_month
    p = tmp_path / "plan.xlsx"; _make_plan_xlsx(str(p))
    plan = parse(str(p), "2026-2027")
    assert [s["school"] for s in schools_of_month(plan, 2026, 9)] == \
        ["北大國高中", "安溪國中"]
    assert schools_of_month(plan, 2026, 7) == []
    # 只有月份的畢典：日期定案前不自動填進任何 6 月，預覽才列得出來
    assert schools_of_month(plan, 2027, 6) == []
    assert [s["school"] for s in schools_of_month(plan, 2027, 6, include_undated=True)] \
        == ["清水國高中畢典"]


def test_api_stats_shape():
    """ministry_stats 需算出 diff/rate；目標未設定時為 None。"""
    grafana.query = lambda sql, fy: [
        {"category": "弟兄會費", "goal": None, "value": 3},
        {"category": "教會聖奉", "goal": 80000, "value": 6400},
    ]
    s = grafana.ministry_stats(2027)
    assert s["弟兄會費"] == {"goal": None, "value": 3, "diff": None, "rate": None}
    assert s["教會聖奉"]["diff"] == -73600
    assert round(s["教會聖奉"]["rate"], 3) == 0.08


# ── LINE 解析 ────────────────────────────────────────────
CURATED = """📖 年度主題與目標 (2026-2027)
• 年度主題：傳道書 10:10
• 支會目標：招募 3位弟兄、2位姊妹
• 贈經目標：3,500本聖經
• 姐妹贈經行程：
 07/18：土城醫院
 08/12(三)：樹林中山路診所
🙏 靈修（實體禱告會時間表）
• 第一週：鶯歌 (週六 16:00)
• 第二週：三峽 (週六 16:00)
• 第三週：樹林 (週六 16:00)
• 第四週：土城 (週日 15:30)
"""


def test_parse_curated_line_summary():
    r = parse_line_text(CURATED, 2026)
    assert r["annual"]["theme"].startswith("傳道書")
    assert r["annual"]["period"] == "2026-2027"
    assert r["annual"]["recruit_target"] == {"brothers": 3, "sisters": 2}
    assert r["annual"]["bible_target"] == 3500
    sched = r["bible_giving"]["schedule"]
    assert {s["date"] for s in sched} == {"2026-07-18", "2026-08-12"}
    assert len(r["prayer_rota"]["weeks"]) == 4


def test_raw_chat_does_not_produce_junk_schedule():
    """原始聊天記錄不該吐出上百筆假排程。"""
    noise = "\n".join(
        "各位基甸勇士：平安！本學期第5次贈經將於2023年11月1日在中正國中舉行，"
        "預計學生中文288本（6箱）茲公佈如下：" for _ in range(50))
    r = parse_line_text(noise, 2026)
    assert len(r.get("bible_giving", {}).get("schedule", [])) == 0


RAW = """2026.06.01 星期一
08:09\t葉忠濶\t圖片
10:58\tCaleb哲斌\t忠濶兄早上好
2026.07.02 星期四
09:00\t王木林\t本週禱告會
"""


def test_split_and_filter_month():
    assert [d for d, _ in split_by_day(RAW)] == ["2026-06-01", "2026-07-02"]
    r = filter_month(RAW, 2026, 6)
    assert r["matched_days"] == 1 and len(r["messages"]) == 2
    assert r["messages"][0]["who"] == "葉忠濶"
    assert r["messages"][1]["text"] == "忠濶兄早上好"
    assert r["range"] == ("2026-06-01", "2026-07-02")


# ── 年度贈經計畫是「固定參考」，不是「上傳當月才有」────────
def _mini_plan_xlsx(path, period="2026-2027"):
    """最小可解析版面：第 1 列年度標題、第 2 列欄位標頭、第 3 列起資料。"""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "配送計畫"
    ws.cell(1, 1, f"{period}年度學校(2026/6/1-2027/5/31)贈經計畫規劃")
    for c, h in enumerate(["月份", "No", "贈經日期", "項目", "學校",
                           "星期", "時間", "贈經數"], start=1):
        ws.cell(2, c, h)
    ws.cell(3, 1, "9月"); ws.cell(3, 2, 1); ws.cell(3, 3, "9月2日")
    ws.cell(3, 4, "國高中"); ws.cell(3, 5, "北大國高中"); ws.cell(3, 7, "15:30")
    wb.save(str(path))


def test_pinned_plan_auto_loads_into_each_month(tmp_path, monkeypatch):
    """隔月產生報告時，-6 仍要吃得到同一份年度計畫（不必重新上傳）。"""
    import importlib
    plan_dir = tmp_path / "贈經計畫"
    plan_dir.mkdir()
    _mini_plan_xlsx(plan_dir / "2026-2027_聖經配送計畫.xlsx")
    monkeypatch.setenv("GIDEONS_BASE_DIR", str(tmp_path / "海山支會"))
    monkeypatch.setenv("GIDEONS_PLAN_DIR", str(plan_dir))
    monkeypatch.setenv("STORAGE", "local")
    monkeypatch.delenv("VERCEL", raising=False)
    import app as A
    A = importlib.reload(A)

    meta = A._meta(2026, 9)                 # 上傳後的「隔月」
    model = A.store.load_model(meta)
    assert not model.get("distribution_plan")
    note = A._ensure_plan_in_model(meta, model)
    assert "自動帶入" in note
    assert model["distribution_plan"]["schools"][0]["school"] == "北大國高中"
    # 已帶入後不重複寫入
    assert A._ensure_plan_in_model(meta, A.store.load_model(meta)) is None


def test_undated_sessions_never_autofill():
    """「06月　日」畢典日期未定：預設不進任何月份，預覽才列出來。"""
    from engine.parse_plan import schools_of_month
    plan = {"schools": [
        {"school": "鳳鳴國中", "date": "2027-04-14", "month": 4, "date_raw": "4月14日"},
        {"school": "清水國高中畢典", "date": None, "month": 6, "date_raw": "06月    日"},
    ]}
    assert schools_of_month(plan, 2026, 6) == []          # 不自動填
    assert schools_of_month(plan, 2027, 6) == []
    assert len(schools_of_month(plan, 2026, 6, include_undated=True)) == 1
    assert [s["school"] for s in schools_of_month(plan, 2027, 4)] == ["鳳鳴國中"]




# ── 固定範本（雲端沒有 Drive 之後唯一的範本來源）────────────
def _fake_templates(d, roc_year=115, month=7):
    d.mkdir(parents=True, exist_ok=True)
    for suffix in ("-1議程", "-2事工成果統計表", "-A.行事曆2026-2027北區"):
        (d / f"月例會議程{roc_year}年{month}月{suffix}.docx").write_bytes(b"PK\x03\x04")
    return d


def test_template_month_read_from_filenames(tmp_path):
    """範本月份一律以檔名為準（民國 115年7月 → 西元 2026, 7）。"""
    from engine.storage import template_month_of
    assert template_month_of(_fake_templates(tmp_path / "t")) == (2026, 7)
    empty = tmp_path / "t2"
    empty.mkdir()
    assert template_month_of(empty) is None


def test_template_falls_back_to_fixed_when_month_folder_missing(tmp_path):
    from engine.dates import build_meta
    from engine.storage import LocalStorage
    base = tmp_path / "海山支會"
    fixed = _fake_templates(tmp_path / "templates")
    s = LocalStorage(base, [tmp_path / "plans"], fixed)
    meta = build_meta(2026, 9)                      # 上月＝2026年8月，資料夾不存在
    assert s.month_template_dir(meta) is None
    assert s.template_dir(meta) == fixed

    real = base / "2026年8月月例會"                  # 有上月資料夾時它優先
    _fake_templates(real, 115, 8)
    assert s.template_dir(meta) == real


def test_fixed_template_rebuilds_meta_month(tmp_path, monkeypatch):
    """用固定範本時 meta.prev_* 必須指向範本的月份。

    否則 update_dates() 要找的是「8月」而範本寫的是「7月」，一個字都換不到，
    產出會安靜地留在舊日期——這比報錯更難發現。
    """
    import importlib
    fixed = _fake_templates(tmp_path / "templates")
    monkeypatch.setenv("GIDEONS_BASE_DIR", str(tmp_path / "海山支會"))
    monkeypatch.setenv("GIDEONS_TEMPLATE_DIR", str(fixed))
    monkeypatch.setenv("GIDEONS_PLAN_DIR", str(tmp_path / "plans"))
    monkeypatch.delenv("VERCEL", raising=False)
    import app as A
    A = importlib.reload(A)

    meta, tpl, label = A.resolve_template(2026, 9)   # 報告 9 月，範本卻是 7 月
    assert tpl == fixed
    assert (meta.prev_year, meta.prev_month) == (2026, 7)
    assert meta.prev_meeting_date == "2026-07-26"    # 第四個禮拜天，換日期用得到
    assert "固定範本" in label
